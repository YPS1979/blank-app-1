import streamlit as st

st.title("🎈 My new app")
st.write(
    "Let's start building! For help and inspiration, head over to [docs.streamlit.io](https://docs.streamlit.io/)."
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Função para calcular o índice de impacto social
def calcular_iis(pontuacoes, pesos):
    return sum(p * w for p, w in zip(pontuacoes, pesos))

# Pesos para cada parâmetro
pesos = [0.10, 0.15, 0.10, 0.10, 0.15, 0.10, 0.10, 0.10, 0.05, 0.05]

# Título da aplicação
st.title("Índice de Impacto Social")

# Perguntas para cada parâmetro
unidade = st.text_input("Unidade/Membro")
p1 = st.slider("P1 - Disponibilidade", 1, 5, 3)
p2 = st.slider("P2 - Indicadores Sociais", 1, 5, 3)
p3 = st.slider("P3 - Planejamento Estratégico", 1, 5, 3)
p4 = st.slider("P4 - Integração", 1, 5, 3)
p5 = st.slider("P5 - Resultados Jurídicos", 1, 5, 3)
p6 = st.slider("P6 - Participação", 1, 5, 3)
p7 = st.slider("P7 - Extrajudicial", 1, 5, 3)
p8 = st.slider("P8 - Coletiva", 1, 5, 3)
p9 = st.slider("P9 - Transformação Social", 1, 5, 3)
p10 = st.slider("P10 - Função e Tempo", 1, 5, 3)

# Botão para calcular o IIS
if st.button("Calcular IIS"):
    pontuacoes = [p1, p2, p3, p4, p5, p6, p7, p8, p9, p10]
    iis = calcular_iis(pontuacoes, pesos)
    st.write(f"Pontuação Total (IIS) para {unidade}: {iis:.2f}")

# Salvar os dados em uma planilha Excel
if st.button("Salvar Dados"):
    data = {
        "Unidade/Membro": [unidade],
        "P1 - Disponibilidade": [p1],
        "P2 - Indicadores Sociais": [p2],
        "P3 - Planejamento Estratégico": [p3],
        "P4 - Integração": [p4],
        "P5 - Resultados Jurídicos": [p5],
        "P6 - Participação": [p6],
        "P7 - Extrajudicial": [p7],
        "P8 - Coletiva": [p8],
        "P9 - Transformação Social": [p9],
        "P10 - Função e Tempo": [p10],
        "Pontuação Total (IIS)": [iis]
    }
    df = pd.DataFrame(data)
    
    # Verificar se o arquivo Excel já existe
    excel_file = 'dados_impacto_social.xlsx'
    try:
        book = load_workbook(excel_file)
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            for sheetname in writer.sheets:
                df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)
    except FileNotFoundError:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
    
    st.success("Dados salvos com sucesso!")
)
